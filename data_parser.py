from rich.console import Console
import matplotlib.pyplot as plt
from pyfiglet import Figlet
from openpyxl import Workbook
from time import sleep

def hexAdder(hexs: list):
    """
    The function takes a list of hexadecimal numbers, adds them together, and returns the sum in
    hexadecimal format.
    
    :param hexs: The parameter "hexs" is a list of hexadecimal numbers that need to be added together
    :type hexs: list
    :return: The function `hexAdder` returns a string that represents the sum of all hexadecimal numbers
    in the input list `hexs`. The string is in uppercase and does not contain the prefix "0x".
    """  # noqa: E501
    temp = 0
    for hexx in hexs:
        temp += int(hexx, 16)
    result = hex(temp)
    return result.replace("0x", "").upper()


def checksumGenerator(added_hex):
    """
    The function generates a checksum in hexadecimal format based on the input value.
    
    :param added_hex: I'm sorry, but the parameter added_hex is not provided in the code snippet you
    shared. Can you please provide the value of added_hex so that I can assist you better?
    :return: a checksum value in hexadecimal format, which is calculated by subtracting the input
    parameter (in hexadecimal format) from FF (also in hexadecimal format) and then converting the
    result back to hexadecimal format. The returned value is in uppercase and does not contain the "0x"
    prefix.
    """  # noqa: E501
    checksum = hex(int('FF', 16) - int(added_hex, 16))
    return checksum.replace("0x", "").upper()

def checksumChecker(hexs: list, checksum):
    """
    --> Calculate the checksum of an API frame

        • Add all bytes of the packet, excluding the start delimiter 0x7E and the length (the second and third bytes).
        • From the result, keep only the lowest 8 bits.
        • Subtract this quantity from 0xFF.

    --> To calculate the checksum for the given frame:
    | Start Delimiter |     Length      | Frame type | 	                        Data	                        | Checksum |
    |       7E  	  |      00 0F      |  	  17     | 	01 	00 	13 	A2 	00 	40 	AD 	14 	2E 	FF 	FE 	02 	44 	42 	|     -    |

        • Add all bytes excluding the start delimiter and the length: 17 + 01 + 00 + 13 + A2 + 00 + 40 + AD + 14 + 2E + FF + FE+ 02 + 44 + 42 = 481
        • From the result, keep only the lowest 8 bits: 81.
        • Subtract that result from 0xFF: FF - 81 = 7E

    In this example, 0x7E is the checksum of the frame.

    --> Checksum Verification

    | Start Delimiter |     Length      | Frame type | 	                        Data	                        | Checksum |
    |       7E  	  |      00 0F      |  	  17     | 	01 	00 	13 	A2 	00 	40 	AD 	14 	2E 	FF 	FE 	02 	44 	42 	|    7E    |

        • Add all data bytes and the checksum: 17 + 01 + 00 + 13 + A2 + 00 + 40 + AD + 14 + 2E + FF + FE + 02 + 44 + 42 + 7E = 4FF
        • Since the last two far right digits of 4FF are FF, the checksum is correct.
    """  # noqa: E501
    hexs.append(checksum)
    result = hexAdder(hexs)

    return True if result[-2:] == 'FF' else False


def recordFragmenter(filenames: list):
    """
    The function takes a list of filenames, reads the files, and returns a list of the third column of
    records where the second column is "RECV".
    
    :param filenames: a list of file names (strings) that contain database records
    :type filenames: list
    :return: The function `recordFragmenter` returns a list of lists. Each inner list contains the
    values of the fourth column of the records in the input files where the third column is equal to
    "RECV". The outer list contains one inner list for each input file.
    """  # noqa: E501
    result_list = []
    for filename in filenames:
        with open(filename, "r", encoding="utf-8") as dbm_record:
            records = dbm_record.readlines()
            results = []
            for record in records[2:]:
                fragmented_record = record.split(",")
                if fragmented_record[2] == "RECV":
                    results.append(fragmented_record[3])
        result_list.append(results)
    return result_list

def hexParser(received_hex) -> dict:
    """
    The function `hexParser` takes a hexadecimal string as input, parses it into various fields,
    calculates the checksum, and returns a dictionary containing the parsed fields and checksum
    information.
    
    :param received_hex: The input parameter received_hex is a string representing a hexadecimal message
    received from a device. The function hexParser parses this message and returns a dictionary
    containing various fields extracted from the message
    :return: The function `hexParser` is returning a dictionary containing various fields parsed from
    the input `received_hex` string. These fields include the start delimiter, length, frame type, frame
    ID, destination address (both 64-bit and 16-bit), AT command, status, dBm response, checksum,
    checksum added, checksum calculated, and a list of the hexadecimal values in the input string.
    """  # noqa: E501
    received_hex = received_hex.replace("\n", "")
    data = received_hex[6:-2]
    hexs = []
    
    for i in range(0, len(data) - 1, 2):
        hexs.append(data[i:i+2])
        
    """The start delimiter is the first byte of a frame consisting of a special
    sequence of bits that indicate the beginning of a data frame.Its value is
    always 0x7E. This allows for easy detection of a new incoming frame."""
    start_delimeter = received_hex[:2]
    """The length field specifies the total number of bytes included in the frame
    data field. Its two-byte value excludes the start delimiter, the length, and
    the checksum. Converted to decimal."""
    length = int(received_hex[2:6], 16) 
    frame_type = received_hex[6:8]
    frame_id = received_hex[8:10]
    destination_address_64bit = received_hex[10:26]
    destination_address_16bit = received_hex[26:30]
    at_command = received_hex[30:34]
    status = received_hex[34:36]
    dbm_response = str(int(received_hex[36:38], 16))
    checksum = received_hex[-2:]
    checksum_added = hexAdder(hexs)[-2:]
    checksum_calculated = checksumGenerator(checksum_added),
    hex_list = hexs

    received_hex_parsed_dict = {
        "start_delimeter": start_delimeter,
        "length": length,
        "frame_type": frame_type,
        "frame_id": frame_id,
        "destination_address_64bit": destination_address_64bit,
        "destination_address_16bit": destination_address_16bit,
        "at_command": at_command,
        "status": status,
        "dbm_response": dbm_response,
        "checksum": checksum,
        "checksum_added": checksum_added,
        "checksum_calculated": checksum_calculated,
        "hex_list": hex_list
    }

    return received_hex_parsed_dict

def writeValues(results):
    """
    The function writes values to an Excel file based on the results of a data parsing process,
    including calculating means and handling checksum errors.
    
    :param results: The `results` parameter is a list of lists, where each inner list contains data from
    a single measurement. Each measurement consists of multiple data points, which are represented as
    strings. The `writeValues` function parses this data and writes it to an Excel file, with each
    column representing a different measurement
    """  # noqa: E501
    console = Console()
    with console.status("[bold yellow]Parsing data..."):
        global means
        means = []
        wb = Workbook()
        sheet = wb.active
        meter = 10
        column = 2

        for k in range(1, 16):
                sheet.cell(row=1, column=k+1).value = meter
                meter += 10
                
        for result in results:
            row = 2
            temp = 0
            
            for i in range(1, len(result)+1):
                sheet.cell(row=i+1, column=1).value = i
            sheet.cell(row=len(result)+2, column=1).value = "Means"
            
            for data in result:
                if not checksumChecker(hexParser(data)["hex_list"], hexParser(data)['checksum']):  # noqa: E501
                    console.log("[bold][red]Checksum error, skipping data[/red]")
                    dbm = 0
                    sheet.cell(row = row, column = column).value = dbm
                    temp += int(dbm)
                    row += 1
                else:
                    dbm = hexParser(data)["dbm_response"]
                    sheet.cell(row = row, column = column).value = -int(dbm)
                    temp += int(dbm)
                    row += 1
            sheet.cell(row = len(result)+2, column = column).value = -float(temp/len(result))  # noqa: E501
            means.append(-float(temp/len(result)))
            column += 1
            console.log("[green]Parsing data finished [/green]")
            sleep(1.5)
        console.log('[bold][blue]Done!')
        wb.save("results.xlsx")

def plotGraph(total_log: int):
    """
    The function plots a graph of RSSI values against distance.
    
    :param total_log: The total number of logs or measurements taken for the RSSI values
    :type total_log: int
    """
    meters = [meter for meter in range(10,(total_log+1)*10,10)]
    plt.plot(meters, means, marker = "o", color = "blue")
    plt.suptitle('Means of RSSI Values', fontsize=16)
    plt.subplots_adjust(left= 0.152)
    plt.ylabel("RSSI Value (dBm)")
    plt.xlabel("Distance (m)")
    # plt.show() # For debugging purposes
    plt.savefig("Means Graph.png")


f = Figlet(font='slant')
print(f.renderText('Hex Parser'))
console = Console()
console.print("[bold]Coded by: [red]Æthereal[/red][/bold]")
console.print("[bold][blue][?] How many log files do you want to parse?: [/blue][/bold]", end="")  # noqa: E501
total_log = int(input()) 
console.print(f"[bold][yellow][i]Parsing {total_log} log files...[/i][/yellow][/bold]")
results = recordFragmenter([str(x) + ".log" for x in range(1, total_log+1)])
writeValues(results)
plotGraph(total_log)

